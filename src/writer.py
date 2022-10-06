import os

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font



class Writer:
    def __init__(self):
        pass

    @staticmethod
    def df_to_csv(normal_df, path):
        # path - путь сохранения файла (в этот путь уже должно быть добавлено наименование файла)
        file_name = path.split("\\")[-1]
        print("++++++++++++++++++++++++++")
        print(f'Сохраняю файл {file_name}')
        print("++++++++++++++++++++++++++")
        print(" ")
        normal_df.to_csv(path, index=False)

    @staticmethod
    def multiple_df_to_csv(multiple_df, src_path):
        # path - путь сохранения файла в котором НЕ прописанно наименование файла
        file_names = list(multiple_df.keys())

        for name in file_names:
            print("+++++++++++++++++++++++")
            print(f"Сохраняю файл 'отделение-{name}'")
            print("+++++++++++++++++++++++")
            print(" ")
            if f"отделение-{name}.csv" in file_names:
                current_data = pd.read_csv(rf"{src_path}/trends/sorted_by_department/отделение-{name}.csv", low_memory=False)
                new_data = pd.concat([multiple_df, current_data], axis=0)
                new_data.to_csv(rf"{src_path}/trends/sorted_by_department/отделение-{name}.csv", index=False)
                print('++++++++++++++++++++++++++++++')
                print(f'Произведено слияние файлов по отделению-{name}')
                print('++++++++++++++++++++++++++++++')
                print(' ')
            else:
                multiple_df[name].to_csv(rf"{src_path}/trends/sorted_by_department/отделение-{name}.csv", index=False)


    @staticmethod
    def multiple_df_to_excel(data_frame, path):
        # path - путь сохранения файла в котором уже прописанно наименование файла
        writer = pd.ExcelWriter(path, engine='openpyxl')
        workbook = writer.book
        days = list(data_frame.keys())

        for day in days:
            data_frame[day].to_excel(writer, sheet_name=f'{day}', startrow=0, startcol=0, index=False)
            sht = workbook[f'{day}']
            for row in sht['A1:MS300']:
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
            column = 1
            while column < 200:
                i = get_column_letter(column)
                sht.column_dimensions[i].width = 20
                column += 1
        workbook.save(path)
        file_name = path.split('\\')[-1]
        print("++++++++++++++++++++++++++++++++++")
        print(f'Сохраняю в формате .xlsx файл {file_name}')
        print("++++++++++++++++++++++++++++++++++")
        print(" ")
